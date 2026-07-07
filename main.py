import os
import sys
import datetime

sys.stdout.reconfigure(encoding='utf-8')

from fastapi import FastAPI, Depends, HTTPException, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from pydantic import BaseModel
import pandas as pd
import openpyxl
import io

import models
import database
import hooks
import ai_agent

# Initialize FastAPI App
app = FastAPI(title="Real-time Task & Budget Management System")

# CORS Middleware (allows Streamlit frontend to interact with the API)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Pydantic schemas for request validation
class TaskProgressUpdate(BaseModel):
    tien_do: float
    trang_thai: str
    dieu_kien_ghi_nhan: Optional[str] = ""

class SpendingCreate(BaseModel):
    task_id: int
    so_tien_chi: float
    nguoi_cap_nhat: str
    chung_tu_kem_theo: Optional[str] = ""
    trang_thai_duyet: Optional[str] = "Approved"

class AIUpdateInput(BaseModel):
    text: str
    api_key: Optional[str] = None

# WebSocket Hub
class ConnectionManager:
    def __init__(self):
        self.active_connections: List[WebSocket] = []

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)
        print(f"WebSocket client connected. Total: {len(self.active_connections)}")

    def disconnect(self, websocket: WebSocket):
        if websocket in self.active_connections:
            self.active_connections.remove(websocket)
            print(f"WebSocket client disconnected. Total: {len(self.active_connections)}")

    async def broadcast(self, message: dict):
        for connection in self.active_connections:
            try:
                await connection.send_json(message)
            except Exception as e:
                # Connection might be closed, we will clean it up later or handle gracefully
                print(f"Error sending message to WebSocket client: {e}")

manager = ConnectionManager()

@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await manager.connect(websocket)
    try:
        while True:
            # Keep connection alive, listen for messages if needed
            data = await websocket.receive_text()
            # Echo back or ignore client messages
            await websocket.send_json({"type": "ping", "message": "alive"})
    except WebSocketDisconnect:
        manager.disconnect(websocket)
    except Exception as e:
        print(f"WebSocket error: {e}")
        manager.disconnect(websocket)

# ----------------- API Endpoints -----------------

import re
import calendar

def parse_date(date_str):
    if not date_str:
        return None
    date_str = str(date_str).strip()
    
    # 2026-06-30
    match1 = re.match(r'^(\d{4})-(\d{2})-(\d{2})$', date_str)
    if match1:
        try:
            return datetime.date(int(match1.group(1)), int(match1.group(2)), int(match1.group(3)))
        except ValueError:
            pass
            
    # Tháng 06/2026
    match2 = re.match(r'^(?:Tháng|tháng|Th\s*)\s*(\d{2})/(\d{4})$', date_str)
    if match2:
        try:
            month = int(match2.group(1))
            year = int(match2.group(2))
            last_day = calendar.monthrange(year, month)[1]
            return datetime.date(year, month, last_day)
        except Exception:
            pass
            
    # 06/2026
    match3 = re.match(r'^(\d{2})/(\d{4})$', date_str)
    if match3:
        try:
            month = int(match3.group(1))
            year = int(match3.group(2))
            last_day = calendar.monthrange(year, month)[1]
            return datetime.date(year, month, last_day)
        except Exception:
            pass
            
    return None

def format_date(d, original_str=None):
    if not d:
        return original_str or "Tháng 06/2026"
    last_day = calendar.monthrange(d.year, d.month)[1]
    if d.day == last_day:
        return f"Tháng {d.month:02d}/{d.year}"
    return d.strftime("%Y-%m-%d")

def compute_rolled_up_deadlines(all_tasks):
    computed = {}
    
    # 1. Parse all tasks' deadlines first
    for t in all_tasks:
        computed[t.stt] = parse_date(t.thoi_han_hoan_thanh)
        
    # 2. Roll up hierarchy (post-order-like traversal via prefix check)
    for t in all_tasks:
        descendant_dates = []
        for other in all_tasks:
            # Check if other is a descendant of t (starts with t.stt + ".")
            if other.stt.startswith(t.stt + "."):
                d_date = parse_date(other.thoi_han_hoan_thanh)
                if d_date:
                    descendant_dates.append(d_date)
                    
        if descendant_dates:
            # Add own date if it exists
            own_date = parse_date(t.thoi_han_hoan_thanh)
            if own_date:
                descendant_dates.append(own_date)
            max_date = max(descendant_dates)
            computed[t.stt] = max_date
            
    return computed

@app.get("/api/project")
def get_project_info(db: Session = Depends(database.get_db)):
    project = db.query(models.Project).filter(models.Project.id == 1).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return {
        "ten_du_an": project.ten_du_an,
        "tong_ngan_sach": project.tong_ngan_sach
    }

@app.get("/api/phases")
def get_phases(db: Session = Depends(database.get_db)):
    phases = db.query(models.Phase).all()
    return phases

@app.get("/api/tasks")
def get_tasks(
    phase_id: Optional[int] = None,
    level: Optional[int] = None,
    search: Optional[str] = None,
    db: Session = Depends(database.get_db)
):
    # Fetch all tasks to compute hierarchy rollups properly
    query_all = db.query(models.Task)
    if phase_id:
        query_all = query_all.filter(models.Task.phase_id == phase_id)
    all_tasks = query_all.all()
    
    # Compute rolled up deadlines
    rolled_up_deadlines = compute_rolled_up_deadlines(all_tasks)
    
    # Now build the filtered list
    result = []
    for t in all_tasks:
        dot_count = t.stt.count('.')
        t_level = dot_count + 1
        
        # Level filter
        if level is not None and t_level != level:
            continue
            
        # Search filter
        if search:
            s_lower = search.lower()
            in_name = s_lower in (t.ten_cong_viec or "").lower()
            in_wbs = s_lower in (t.ma_ngan_sach or "").lower()
            in_stt = s_lower in (t.stt or "").lower()
            if not (in_name or in_wbs or in_stt):
                continue
                
        # Get rolled up deadline or keep original if none
        rolled_date = rolled_up_deadlines.get(t.stt)
        final_deadline = format_date(rolled_date, t.thoi_han_hoan_thanh)
        
        result.append({
            "id": t.id,
            "ma_ngan_sach": t.ma_ngan_sach,
            "stt": t.stt,
            "level": t_level,
            "phase_id": t.phase_id,
            "ten_cong_viec": t.ten_cong_viec,
            "phong_ban_thuc_hien": t.phong_ban_thuc_hien,
            "co_quan_giai_quyet": t.co_quan_giai_quyet,
            "ho_so_dau_ra": t.ho_so_dau_ra,
            "dieu_kien_ghi_nhan": t.dieu_kien_ghi_nhan,
            "thoi_han_hoan_thanh": final_deadline,
            "tien_do": t.tien_do,
            "trang_thai": t.trang_thai,
            "ke_hoach_tuan": t.ke_hoach_tuan,
            "ket_qua_tuan": t.ket_qua_tuan,
            "vuong_mac_tuan": t.vuong_mac_tuan,
            "cach_giai_quyet": t.cach_giai_quyet,
            "duyet_tuan": t.duyet_tuan,
            "budget": {
                "ngan_sach_tong": t.budget.ngan_sach_tong if t.budget else 0.0,
                "is_locked": t.budget.is_locked if t.budget else False
            } if t.budget else None
        })
        
    return result

@app.get("/api/tasks/export")
def export_tasks_to_excel(db: Session = Depends(database.get_db)):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import DataBarRule
    
    tasks = db.query(models.Task).all()
    rolled_up_deadlines = compute_rolled_up_deadlines(tasks)
    
    # Sort them by STT hierarchically
    tasks = sorted(tasks, key=lambda x: [int(i) if i.isdigit() else 999 for i in x.stt.split('.')])
    
    rows = []
    for t in tasks:
        dot_count = t.stt.count('.')
        t_level = dot_count + 1
        budget_val = t.budget.ngan_sach_tong if t.budget else 0.0
        
        # Indent task name to represent level visually in Excel
        indented_name = ("    " * (t_level - 1)) + t.ten_cong_viec
        
        # Get rolled up deadline or keep original if none
        rolled_date = rolled_up_deadlines.get(t.stt)
        final_deadline = format_date(rolled_date, t.thoi_han_hoan_thanh)
        
        rows.append({
            "STT": t.stt,
            "Cấp": t_level,
            "Mã Ngân Sách (WBS)": t.ma_ngan_sach,
            "Nội dung công việc": indented_name,
            "Phòng ban thực hiện": t.phong_ban_thuc_hien or "-",
            "Hồ sơ đầu ra": t.ho_so_dau_ra or "-",
            "Thời hạn hoàn thành": final_deadline or "-",
            "Điều kiện ghi nhận kết quả": t.dieu_kien_ghi_nhan or "-",
            "Ngân sách tổng (Trđ)": budget_val,
            "Tiến độ (%)": (t.tien_do / 100.0) if t.tien_do else 0.0,
            "Kế hoạch tuần": t.ke_hoach_tuan or "-",
            "Kết quả tuần": t.ket_qua_tuan or "-",
            "Vướng mắc tuần": t.vuong_mac_tuan or "-",
            "Giải quyết của CBQL/Phòng ban": t.cach_giai_quyet or "-",
            "Duyệt tuần": t.duyet_tuan or "Chưa duyệt",
            "Trạng thái": t.trang_thai
        })
        
    df = pd.DataFrame(rows)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Ke_hoach_cong_viec")
        
        workbook = writer.book
        worksheet = writer.sheets["Ke_hoach_cong_viec"]
        
        # Freeze header row
        worksheet.freeze_panes = 'A2'
        
        # Border styles
        thin_border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )
        
        # Color palettes
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        level1_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Blue
        level2_fill = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid") # Sky blue 100
        level3_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") # Slate 50
        
        # Font configurations
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        level1_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        level2_font = Font(name="Segoe UI", size=10, bold=True, color="0369A1")
        level3_font = Font(name="Segoe UI", size=10, bold=False, color="1E293B")
        standard_font = Font(name="Segoe UI", size=10, bold=False, color="334155")
        
        # Style Header
        worksheet.row_dimensions[1].height = 28
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            
        # Style Rows
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(tasks) + 1), start=2):
            task_level = row[1].value  # "Cấp" is Column B (index 1)
            worksheet.row_dimensions[row_idx].height = 22
            
            # Select level styling
            if task_level == 1:
                row_fill = level1_fill
                row_font = level1_font
            elif task_level == 2:
                row_fill = level2_fill
                row_font = level2_font
            elif task_level == 3:
                row_fill = level3_fill
                row_font = level3_font
            else:
                row_fill = None
                row_font = standard_font
                
            for col_idx, cell in enumerate(row):
                if row_fill:
                    cell.fill = row_fill
                cell.font = row_font
                cell.border = thin_border
                
                # Alignments
                if col_idx in [0, 1, 2, 6, 14, 15]:  # STT, Cấp, WBS, Thời hạn, Duyệt tuần, Trạng thái
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_idx in [8]:  # Ngân sách
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_idx in [9]:  # Tiến độ
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
            # Number formats
            row[8].number_format = '#,##0.00" Trđ"'
            row[9].number_format = '0.0%'
            
        # Add Progress Data Bar (Green color 10B981)
        data_bar_rule = DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1.0, color="10B981", showValue=True)
        worksheet.conditional_formatting.add(f"J2:J{len(tasks)+1}", data_bar_rule)
        
        # Auto-adjust column widths
        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            # Give column D (indented name) extra width
            if col_letter == 'D':
                worksheet.column_dimensions[col_letter].width = max(max_len + 8, 30)
            else:
                worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    output.seek(0)
    
    headers = {
        'Content-Disposition': 'attachment; filename="VenSongVinh_WBS_KeHoach.xlsx"'
    }
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

@app.get("/api/tasks/{task_id}")
def get_task(task_id: int, db: Session = Depends(database.get_db)):
    task = db.query(models.Task).filter(models.Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
        
    # Compute rolled up deadline
    all_tasks = db.query(models.Task).filter(models.Task.phase_id == task.phase_id).all()
    rolled_up_deadlines = compute_rolled_up_deadlines(all_tasks)
    rolled_date = rolled_up_deadlines.get(task.stt)
    final_deadline = format_date(rolled_date, task.thoi_han_hoan_thanh)
    
    return {
        "id": task.id,
        "ma_ngan_sach": task.ma_ngan_sach,
        "stt": task.stt,
        "phase_id": task.phase_id,
        "ten_cong_viec": task.ten_cong_viec,
        "phong_ban_thuc_hien": task.phong_ban_thuc_hien,
        "co_quan_giai_quyet": task.co_quan_giai_quyet,
        "ho_so_dau_ra": task.ho_so_dau_ra,
        "dieu_kien_ghi_nhan": task.dieu_kien_ghi_nhan,
        "thoi_han_hoan_thanh": final_deadline,
        "tien_do": task.tien_do,
        "trang_thai": task.trang_thai,
        "ke_hoach_tuan": task.ke_hoach_tuan,
        "ket_qua_tuan": task.ket_qua_tuan,
        "vuong_mac_tuan": task.vuong_mac_tuan,
        "cach_giai_quyet": task.cach_giai_quyet,
        "duyet_tuan": task.duyet_tuan,
        "budget": {
            "ngan_sach_tong": task.budget.ngan_sach_tong if task.budget else 0.0,
            "is_locked": task.budget.is_locked if task.budget else False
        } if task.budget else None
    }

@app.put("/api/tasks/{task_id}/progress")
async def update_task_progress(
    task_id: int, 
    update_data: TaskProgressUpdate, 
    user_role: str = "PM", # Can be Admin or PM
    db: Session = Depends(database.get_db)
):
    task = db.query(models.Task).filter(models.Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
        
    # Layer 4: Phase Gate Loop check
    try:
        hooks.execute_phase_gate_loop(db, task_id, update_data.trang_thai)
    except hooks.PhaseGateException as e:
        raise HTTPException(status_code=400, detail=str(e))

    # Update database
    task.tien_do = update_data.tien_do
    task.trang_thai = update_data.trang_thai
    if update_data.dieu_kien_ghi_nhan:
        task.dieu_kien_ghi_nhan = update_data.dieu_kien_ghi_nhan

    db.add(task)
    db.commit()

    # Trigger real-time WebSocket broadcast
    update_msg = {
        "type": "task_update",
        "task_id": task.id,
        "ma_ngan_sach": task.ma_ngan_sach,
        "stt": task.stt,
        "ten_cong_viec": task.ten_cong_viec,
        "tien_do": task.tien_do,
        "trang_thai": task.trang_thai,
        "dieu_kien_ghi_nhan": task.dieu_kien_ghi_nhan
    }
    await manager.broadcast(update_msg)

    return {"status": "success", "task": update_msg}

@app.post("/api/spending")
async def add_spending(
    spending: SpendingCreate, 
    user_role: str = "PM", 
    db: Session = Depends(database.get_db)
):
    task = db.query(models.Task).filter(models.Task.id == spending.task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    # Layer 4: Hard Budget Gate check
    # If the spending is Approved, check the budget limits
    if spending.trang_thai_duyet == "Approved":
        try:
            hooks.execute_hard_budget_gate(db, spending.task_id, spending.so_tien_chi)
        except hooks.BudgetExceededException as e:
            # We notify C-level dashboard about red alert
            alert_msg = {
                "type": "red_alert",
                "message": str(e),
                "ma_ngan_sach": task.ma_ngan_sach,
                "task_id": task.id
            }
            await manager.broadcast(alert_msg)
            raise HTTPException(status_code=400, detail=str(e))

    # Save spending record
    new_spend = models.ActualSpending(
        task_id=spending.task_id,
        so_tien_chi=spending.so_tien_chi,
        ngay_chi=datetime.date.today(),
        nguoi_cap_nhat=spending.nguoi_cap_nhat,
        chung_tu_kem_theo=spending.chung_tu_kem_theo,
        trang_thai_duyet=spending.trang_thai_duyet
    )
    db.add(new_spend)
    db.commit()

    # Broadcast spending update
    spending_msg = {
        "type": "spending_update",
        "task_id": task.id,
        "ma_ngan_sach": task.ma_ngan_sach,
        "so_tien_chi": spending.so_tien_chi,
        "nguoi_cap_nhat": spending.nguoi_cap_nhat,
        "trang_thai_duyet": spending.trang_thai_duyet
    }
    await manager.broadcast(spending_msg)

    return {"status": "success", "spending_id": new_spend.id}

@app.get("/api/spending")
def get_spending(db: Session = Depends(database.get_db)):
    spendings = db.query(models.ActualSpending).all()
    result = []
    for s in spendings:
        task = db.query(models.Task).filter(models.Task.id == s.task_id).first()
        result.append({
            "id": s.id,
            "task_id": s.task_id,
            "ma_ngan_sach": task.ma_ngan_sach if task else "N/A",
            "ten_cong_viec": task.ten_cong_viec if task else "N/A",
            "so_tien_chi": s.so_tien_chi,
            "ngay_chi": s.ngay_chi.isoformat(),
            "nguoi_cap_nhat": s.nguoi_cap_nhat,
            "chung_tu_kem_theo": s.chung_tu_kem_theo,
            "trang_thai_duyet": s.trang_thai_duyet
        })
    return result

@app.get("/api/stats")
def get_dashboard_stats(db: Session = Depends(database.get_db)):
    project = db.query(models.Project).filter(models.Project.id == 1).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # Total budget
    total_budget = project.tong_ngan_sach

    # Total actual spending approved
    total_spent = db.query(models.ActualSpending).filter(
        models.ActualSpending.trang_thai_duyet == "Approved"
    ).all()
    total_spent_val = sum(s.so_tien_chi for s in total_spent)

    # Average progress of Level 1 tasks
    lvl_1_tasks = db.query(models.Task).all()
    # Filter Level 1
    lvl_1_tasks = [t for t in lvl_1_tasks if len(t.stt.split('.')) == 1]
    avg_progress = sum(t.tien_do for t in lvl_1_tasks) / len(lvl_1_tasks) if lvl_1_tasks else 0.0

    # Count tasks by status
    all_tasks = db.query(models.Task).all()
    todo_count = sum(1 for t in all_tasks if t.trang_thai == "Todo")
    inprogress_count = sum(1 for t in all_tasks if t.trang_thai == "In-Progress")
    done_count = sum(1 for t in all_tasks if t.trang_thai == "Done")
    delayed_count = sum(1 for t in all_tasks if t.trang_thai == "Delayed")

    # Locked budgets count
    locked_count = db.query(models.Budget).filter(models.Budget.is_locked == True).count()

    return {
        "tong_ngan_sach": total_budget,
        "tong_thuc_chi": total_spent_val,
        "con_lai": total_budget - total_spent_val,
        "tien_do_trung_binh": avg_progress,
        "trang_thai_tasks": {
            "Todo": todo_count,
            "In-Progress": inprogress_count,
            "Done": done_count,
            "Delayed": delayed_count
        },
        "so_wbs_bi_khoa": locked_count
    }

@app.get("/api/s-curve")
def get_s_curve_data(db: Session = Depends(database.get_db)):
    """
    Tạo dữ liệu đường cong S-Curve kế hoạch vs thực tế chi tiêu.
    Các mốc thời gian: 
    - 01/2025, 02/2025, 03/2025, 04/2025
    - Quý 1/2026 (biểu diễn mốc 03/2026)
    - 05/2026, 06/2026
    """
    budgets = db.query(models.Budget).all()
    
    # Kế hoạch chi tiêu cho từng mốc (tính tổng của toàn bộ các task)
    plan_t1_2025 = sum(b.thang_01_2025 for b in budgets)
    plan_t2_2025 = sum(b.thang_02_2025 for b in budgets)
    plan_t3_2025 = sum(b.thang_03_2025 for b in budgets)
    plan_t4_2025 = sum(b.thang_04_2025 for b in budgets)
    plan_q1_2026 = sum(b.quy_1_2026 for b in budgets)
    plan_t5_2026 = sum(b.thang_05_2026 for b in budgets)
    plan_t6_2026 = sum(b.thang_06_2026 for b in budgets)

    # Lũy kế kế hoạch chi tiêu
    cum_plan = []
    
    # 01/2025
    cum_plan.append({"period": "Tháng 01/2025", "value": plan_t1_2025})
    # 02/2025
    cum_plan.append({"period": "Tháng 02/2025", "value": plan_t1_2025 + plan_t2_2025})
    # 03/2025
    cum_plan.append({"period": "Tháng 03/2025", "value": plan_t1_2025 + plan_t2_2025 + plan_t3_2025})
    # 04/2025
    cum_plan.append({"period": "Tháng 04/2025", "value": plan_t1_2025 + plan_t2_2025 + plan_t3_2025 + plan_t4_2025})
    # Quý 1/2026
    cum_plan.append({"period": "Quý 1/2026", "value": plan_t1_2025 + plan_t2_2025 + plan_t3_2025 + plan_t4_2025 + plan_q1_2026})
    # 05/2026
    cum_plan.append({"period": "Tháng 05/2026", "value": plan_t1_2025 + plan_t2_2025 + plan_t3_2025 + plan_t4_2025 + plan_q1_2026 + plan_t5_2026})
    # 06/2026
    cum_plan.append({"period": "Tháng 06/2026", "value": plan_t1_2025 + plan_t2_2025 + plan_t3_2025 + plan_t4_2025 + plan_q1_2026 + plan_t5_2026 + plan_t6_2026})

    # Lũy kế thực chi theo mốc thời gian thực tế nhập trong actual_spending
    spendings = db.query(models.ActualSpending).filter(
        models.ActualSpending.trang_thai_duyet == "Approved"
    ).all()

    # Phân bổ thực chi vào các mốc thời gian tương ứng
    actual_t1_2025 = 0.0
    actual_t2_2025 = 0.0
    actual_t3_2025 = 0.0
    actual_t4_2025 = 0.0
    actual_q1_2026 = 0.0
    actual_t5_2026 = 0.0
    actual_t6_2026 = 0.0

    for s in spendings:
        date = s.ngay_chi
        # Phân chia thời gian giải ngân
        if date.year == 2025:
            if date.month == 1:
                actual_t1_2025 += s.so_tien_chi
            elif date.month == 2:
                actual_t2_2025 += s.so_tien_chi
            elif date.month == 3:
                actual_t3_2025 += s.so_tien_chi
            elif date.month == 4:
                actual_t4_2025 += s.so_tien_chi
            else:
                # Gán tạm vào tháng 4 nếu là các tháng khác của 2025
                actual_t4_2025 += s.so_tien_chi
        elif date.year == 2026:
            if date.month in [1, 2, 3]:
                actual_q1_2026 += s.so_tien_chi
            elif date.month == 5:
                actual_t5_2026 += s.so_tien_chi
            elif date.month == 6:
                actual_t6_2026 += s.so_tien_chi
            else:
                actual_t6_2026 += s.so_tien_chi
        else:
            # Năm 2027+
            actual_t6_2026 += s.so_tien_chi

    cum_actual = []
    # 01/2025
    cum_actual.append({"period": "Tháng 01/2025", "value": actual_t1_2025})
    # 02/2025
    cum_actual.append({"period": "Tháng 02/2025", "value": actual_t1_2025 + actual_t2_2025})
    # 03/2025
    cum_actual.append({"period": "Tháng 03/2025", "value": actual_t1_2025 + actual_t2_2025 + actual_t3_2025})
    # 04/2025
    cum_actual.append({"period": "Tháng 04/2025", "value": actual_t1_2025 + actual_t2_2025 + actual_t3_2025 + actual_t4_2025})
    # Quý 1/2026
    cum_actual.append({"period": "Quý 1/2026", "value": actual_t1_2025 + actual_t2_2025 + actual_t3_2025 + actual_t4_2025 + actual_q1_2026})
    # 05/2026
    cum_actual.append({"period": "Tháng 05/2026", "value": actual_t1_2025 + actual_t2_2025 + actual_t3_2025 + actual_t4_2025 + actual_q1_2026 + actual_t5_2026})
    # 06/2026
    cum_actual.append({"period": "Tháng 06/2026", "value": actual_t1_2025 + actual_t2_2025 + actual_t3_2025 + actual_t4_2025 + actual_q1_2026 + actual_t5_2026 + actual_t6_2026})

    return {
        "plan": cum_plan,
        "actual": cum_actual
    }

@app.post("/api/ai/parse")
async def parse_and_update_task(payload: AIUpdateInput, db: Session = Depends(database.get_db)):
    """
    AI Integration Endpoint
    Nhận văn bản thô -> Gemini bóc tách -> Tự động cập nhật Task thành Done và điền điều kiện ghi nhận
    """
    result = ai_agent.parse_natural_language_update(payload.text, payload.api_key)
    
    ma_ngan_sach = result.get("ma_ngan_sach", "")
    trang_thai = result.get("trang_thai", "Done")
    tien_do = result.get("tien_do", 100.0)
    dieu_kien = result.get("dieu_kien_ghi_nhan", "")

    if not ma_ngan_sach:
        raise HTTPException(status_code=400, detail="AI không thể phát hiện mã ngân sách/STT phù hợp trong văn bản.")

    # Tìm task phù hợp trong DB (kiểm tra cả ma_ngan_sach và stt)
    task = db.query(models.Task).filter(
        (models.Task.ma_ngan_sach == ma_ngan_sach) |
        (models.Task.stt == ma_ngan_sach)
    ).first()

    if not task:
        # Thử tìm theo dạng chứa phần cuối của STT
        task = db.query(models.Task).filter(models.Task.stt.like(f"%{ma_ngan_sach}%")).first()

    if not task:
        return {
            "status": "partial_success",
            "message": f"AI trích xuất thành công nhưng không tìm thấy công việc tương ứng cho mã '{ma_ngan_sach}'.",
            "extracted_data": result
        }

    # Chạy chốt chặn Phase Gate Loop trước khi cập nhật
    try:
        hooks.execute_phase_gate_loop(db, task.id, trang_thai)
    except hooks.PhaseGateException as e:
        raise HTTPException(status_code=400, detail=f"Lỗi Logic Chốt chặn: {str(e)}")

    # Tiến hành cập nhật
    task.tien_do = tien_do
    task.trang_thai = trang_thai
    task.dieu_kien_ghi_nhan = dieu_kien
    
    db.add(task)
    db.commit()

    # Broadcast WebSocket update
    update_msg = {
        "type": "task_update",
        "task_id": task.id,
        "ma_ngan_sach": task.ma_ngan_sach,
        "stt": task.stt,
        "ten_cong_viec": task.ten_cong_viec,
        "tien_do": task.tien_do,
        "trang_thai": task.trang_thai,
        "dieu_kien_ghi_nhan": task.dieu_kien_ghi_nhan
    }
    await manager.broadcast(update_msg)

    return {
        "status": "success",
        "message": f"Đã tự động cập nhật công việc {task.stt} thành công qua {result['method']}.",
        "task": update_msg
    }

@app.get("/api/ai/risk")
def get_ai_risk_assessment(api_key: Optional[str] = None, db: Session = Depends(database.get_db)):
    project = db.query(models.Project).filter(models.Project.id == 1).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    total_spent = db.query(models.ActualSpending).filter(
        models.ActualSpending.trang_thai_duyet == "Approved"
    ).all()
    total_spent_val = sum(s.so_tien_chi for s in total_spent)

    # Lấy thông tin 10 khoản chi gần nhất để AI phân tích
    spending_details = []
    for s in total_spent[-10:]:
        task = db.query(models.Task).filter(models.Task.id == s.task_id).first()
        spending_details.append({
            "ma_ngan_sach": task.ma_ngan_sach if task else "N/A",
            "ten_cong_viec": task.ten_cong_viec if task else "N/A",
            "so_tien_chi": s.so_tien_chi,
            "nguoi_chi": s.nguoi_cap_nhat
        })

    risk_report = ai_agent.evaluate_financial_risk(
        project.tong_ngan_sach,
        total_spent_val,
        spending_details,
        api_key
    )
    
    return {"risk_report": risk_report}

class AIChatRequest(BaseModel):
    question: str
    context: str
    api_key: Optional[str] = None

@app.post("/api/ai/chat")
async def chat_with_gemini(request: AIChatRequest):
    api_key = request.api_key or os.environ.get("GEMINI_API_KEY")
    if not api_key:
        raise HTTPException(status_code=400, detail="Gemini API Key is missing. Please configure it in the sidebar.")
    
    prompt = f"""
Bạn là Chuyên gia Kiến trúc Hệ thống ERP và Phân tích Dữ liệu AI của Dự án Bất động sản Ven Sông Vinh.
Dưới đây là dữ liệu thực tế thời gian thực của dự án:
{request.context}

Câu hỏi của người dùng:
"{request.question}"

Hãy phân tích dữ liệu và trả lời câu hỏi một cách ngắn gọn, súc tích, chuyên nghiệp bằng Tiếng Việt.
"""
    try:
        response_text = ai_agent.generate_generic_text(prompt, api_key=api_key)
        return {"response": response_text}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

class TaskCreate(BaseModel):
    parent_stt: str
    ten_cong_viec: str
    phong_ban_thuc_hien: Optional[str] = "BQLDA"
    co_quan_giai_quyet: Optional[str] = "-"
    ho_so_dau_ra: Optional[str] = "-"
    dieu_kien_ghi_nhan: Optional[str] = "-"
    thoi_han_hoan_thanh: Optional[str] = "Tháng 06/2026"
    ngan_sach: float = 0.0

def recalculate_budgets(db: Session):
    remaining_tasks = db.query(models.Task).all()
    task_by_stt = {t.stt: t for t in remaining_tasks}

    all_stts = {t.stt for t in remaining_tasks}
    parent_tasks = []
    
    for t in remaining_tasks:
        is_parent = any(other.startswith(f"{t.stt}.") for other in all_stts)
        if is_parent:
            parent_tasks.append(t)
            
    for t in parent_tasks:
        if t.budget:
            b = t.budget
            b.ngan_sach_tong = 0.0
            b.kh_2026 = 0.0
            b.quy_1_2026 = 0.0
            b.quy_2_2026 = 0.0
            b.thang_01_2025 = 0.0
            b.thang_02_2025 = 0.0
            b.thang_03_2025 = 0.0
            b.thang_04_2025 = 0.0
            b.thang_05_2026 = 0.0
            b.thang_06_2026 = 0.0
            db.add(b)
    db.flush()

    sorted_tasks = sorted(remaining_tasks, key=lambda t: len(t.stt.split('.')), reverse=True)

    for task in sorted_tasks:
        stt_parts = task.stt.split('.')
        if len(stt_parts) <= 1:
            continue
        
        parent_stt = ".".join(stt_parts[:-1])
        parent_task = task_by_stt.get(parent_stt)

        if parent_task:
            child_budget = task.budget
            parent_budget = parent_task.budget
            
            if child_budget and parent_budget:
                parent_budget.ngan_sach_tong += child_budget.ngan_sach_tong
                parent_budget.kh_2026 += child_budget.kh_2026
                parent_budget.quy_1_2026 += child_budget.quy_1_2026
                parent_budget.quy_2_2026 += child_budget.quy_2_2026
                parent_budget.thang_01_2025 += child_budget.thang_01_2025
                parent_budget.thang_02_2025 += child_budget.thang_02_2025
                parent_budget.thang_03_2025 += child_budget.thang_03_2025
                parent_budget.thang_04_2025 += child_budget.thang_04_2025
                parent_budget.thang_05_2026 += child_budget.thang_05_2026
                parent_budget.thang_06_2026 += child_budget.thang_06_2026
                db.add(parent_budget)
    db.flush()

    level_1_tasks = [t for t in remaining_tasks if len(t.stt.split('.')) == 1]
    project_total = sum(t.budget.ngan_sach_tong for t in level_1_tasks if t.budget)

    project_obj = db.query(models.Project).filter(models.Project.id == 1).first()
    if project_obj:
        project_obj.tong_ngan_sach = project_total
        db.add(project_obj)
    db.flush()

@app.post("/api/tasks")
async def create_task(request: TaskCreate, db: Session = Depends(database.get_db)):
    parent = db.query(models.Task).filter(models.Task.stt == request.parent_stt).first()
    if not parent:
        raise HTTPException(status_code=404, detail=f"Không tìm thấy công việc cha có STT {request.parent_stt}")
        
    parent_parts = parent.stt.split('.')
    parent_level = len(parent_parts)
        
    siblings = db.query(models.Task).filter(models.Task.stt.like(f"{request.parent_stt}.%")).all()
    sibling_indices = []
    for s in siblings:
        parts = s.stt.split('.')
        if len(parts) == parent_level + 1 and parts[-1].isdigit():
            sibling_indices.append(int(parts[-1]))
            
    next_idx = max(sibling_indices) + 1 if sibling_indices else 1
    new_stt = f"{request.parent_stt}.{next_idx}"
    new_wbs = f"{parent.ma_ngan_sach}.{next_idx}"
    
    new_task = models.Task(
        project_id=1,
        ma_ngan_sach=new_wbs,
        stt=new_stt,
        phase_id=parent.phase_id,
        ten_cong_viec=request.ten_cong_viec.strip(),
        phong_ban_thuc_hien=request.phong_ban_thuc_hien.strip(),
        co_quan_giai_quyet=request.co_quan_giai_quyet.strip(),
        ho_so_dau_ra=request.ho_so_dau_ra.strip(),
        dieu_kien_ghi_nhan=request.dieu_kien_ghi_nhan.strip(),
        thoi_han_hoan_thanh=request.thoi_han_hoan_thanh.strip(),
        tien_do=0.0,
        trang_thai="Todo"
    )
    db.add(new_task)
    db.flush()
    
    new_budget = models.Budget(
        task_id=new_task.id,
        ngan_sach_tong=request.ngan_sach,
        kh_2026=request.ngan_sach * 0.40,
        quy_1_2026=request.ngan_sach * 0.15,
        quy_2_2026=request.ngan_sach * 0.15,
        thang_01_2025=request.ngan_sach * 0.05,
        thang_02_2025=request.ngan_sach * 0.05,
        thang_03_2025=request.ngan_sach * 0.05,
        thang_04_2025=request.ngan_sach * 0.05,
        thang_05_2026=request.ngan_sach * 0.05,
        thang_06_2026=request.ngan_sach * 0.05
    )
    db.add(new_budget)
    db.flush()
    
    recalculate_budgets(db)
    db.commit()
    
    update_msg = {
        "type": "task_create",
        "task_id": new_task.id,
        "stt": new_task.stt,
        "ma_ngan_sach": new_task.ma_ngan_sach,
        "ten_cong_viec": new_task.ten_cong_viec,
        "phong_ban_thuc_hien": new_task.phong_ban_thuc_hien,
        "ho_so_dau_ra": new_task.ho_so_dau_ra,
        "thoi_han_hoan_thanh": new_task.thoi_han_hoan_thanh,
        "tien_do": new_task.tien_do,
        "trang_thai": new_task.trang_thai,
        "budget": {
            "ngan_sach_tong": new_budget.ngan_sach_tong,
            "is_locked": False
        }
    }
    await manager.broadcast(update_msg)
    
    return {
        "status": "success",
        "message": f"Đã thêm mới thành công công việc Cấp 3: {new_stt}",
        "task": update_msg
    }

class TaskUpdate(BaseModel):
    ma_ngan_sach: str
    ten_cong_viec: str
    phong_ban_thuc_hien: str
    co_quan_giai_quyet: str
    ho_so_dau_ra: str
    dieu_kien_ghi_nhan: str
    thoi_han_hoan_thanh: str
    tien_do: float
    trang_thai: str
    ngan_sach: float
    ke_hoach_tuan: Optional[str] = ""
    ket_qua_tuan: Optional[str] = ""
    vuong_mac_tuan: Optional[str] = ""
    cach_giai_quyet: Optional[str] = ""
    duyet_tuan: Optional[str] = "Chưa duyệt"

@app.put("/api/tasks/{task_id}")
async def update_task_details(task_id: int, request: TaskUpdate, db: Session = Depends(database.get_db)):
    task = db.query(models.Task).filter(models.Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Không tìm thấy công việc")
        
    if request.ma_ngan_sach != task.ma_ngan_sach:
        existing = db.query(models.Task).filter(models.Task.ma_ngan_sach == request.ma_ngan_sach).first()
        if existing:
            raise HTTPException(status_code=400, detail="Mã Ngân Sách (WBS) này đã tồn tại.")
            
    task.ma_ngan_sach = request.ma_ngan_sach.strip()
    task.ten_cong_viec = request.ten_cong_viec.strip()
    task.phong_ban_thuc_hien = request.phong_ban_thuc_hien.strip()
    task.co_quan_giai_quyet = request.co_quan_giai_quyet.strip()
    task.ho_so_dau_ra = request.ho_so_dau_ra.strip()
    task.dieu_kien_ghi_nhan = request.dieu_kien_ghi_nhan.strip()
    task.thoi_han_hoan_thanh = request.thoi_han_hoan_thanh.strip()
    task.tien_do = request.tien_do
    task.trang_thai = request.trang_thai
    task.ke_hoach_tuan = request.ke_hoach_tuan.strip() if request.ke_hoach_tuan else ""
    task.ket_qua_tuan = request.ket_qua_tuan.strip() if request.ket_qua_tuan else ""
    task.vuong_mac_tuan = request.vuong_mac_tuan.strip() if request.vuong_mac_tuan else ""
    task.cach_giai_quyet = request.cach_giai_quyet.strip() if request.cach_giai_quyet else ""
    task.duyet_tuan = request.duyet_tuan.strip() if request.duyet_tuan else "Chưa duyệt"
    
    if task.budget:
        task.budget.ngan_sach_tong = request.ngan_sach
        task.budget.kh_2026 = request.ngan_sach * 0.40
        task.budget.quy_1_2026 = request.ngan_sach * 0.15
        task.budget.quy_2_2026 = request.ngan_sach * 0.15
        task.budget.thang_01_2025 = request.ngan_sach * 0.05
        task.budget.thang_02_2025 = request.ngan_sach * 0.05
        task.budget.thang_03_2025 = request.ngan_sach * 0.05
        task.budget.thang_04_2025 = request.ngan_sach * 0.05
        task.budget.thang_05_2026 = request.ngan_sach * 0.05
        task.budget.thang_06_2026 = request.ngan_sach * 0.05
    db.add(task)
    db.flush()
    
    recalculate_budgets(db)
    db.commit()
    
    update_msg = {
        "type": "task_update",
        "task_id": task.id,
        "stt": task.stt,
        "ma_ngan_sach": task.ma_ngan_sach,
        "ten_cong_viec": task.ten_cong_viec,
        "phong_ban_thuc_hien": task.phong_ban_thuc_hien,
        "ho_so_dau_ra": task.ho_so_dau_ra,
        "thoi_han_hoan_thanh": task.thoi_han_hoan_thanh,
        "tien_do": task.tien_do,
        "trang_thai": task.trang_thai,
        "ke_hoach_tuan": task.ke_hoach_tuan,
        "ket_qua_tuan": task.ket_qua_tuan,
        "vuong_mac_tuan": task.vuong_mac_tuan,
        "cach_giai_quyet": task.cach_giai_quyet,
        "duyet_tuan": task.duyet_tuan,
        "budget": {
            "ngan_sach_tong": task.budget.ngan_sach_tong if task.budget else 0.0,
            "is_locked": task.budget.is_locked if task.budget else False
        }
    }
    await manager.broadcast(update_msg)
    
    return {"status": "success", "task": update_msg}

@app.delete("/api/tasks/{task_id}")
async def delete_task_route(task_id: int, db: Session = Depends(database.get_db)):
    task = db.query(models.Task).filter(models.Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Không tìm thấy công việc")
        
    stt = task.stt
    children = db.query(models.Task).filter(
        (models.Task.stt == stt) | (models.Task.stt.like(f"{stt}.%"))
    ).all()
    
    deleted_ids = [c.id for c in children]
    for c in children:
        db.delete(c)
    db.flush()
    
    recalculate_budgets(db)
    db.commit()
    
    await manager.broadcast({
        "type": "task_delete",
        "task_ids": deleted_ids
    })
    
    return {"status": "success", "deleted_count": len(children)}

# Serve Frontend static files
# We will create the static directory if not exists
os.makedirs("./static", exist_ok=True)
app.mount("/", StaticFiles(directory="./static", html=True), name="static")

import asyncio

async def monitor_observability_loop():
    await asyncio.sleep(5)  # Chờ ứng dụng khởi động hoàn toàn
    while True:
        db = database.SessionLocal()
        try:
            # 1. Kiểm tra trễ hạn công việc
            # Thời điểm hiện tại là Tháng 7/2026. Do đó các kỳ hạn trong quá khứ nếu
            # có ngân sách kế hoạch lớn hơn 0 mà tiến độ < 100% thì được coi là trễ hạn.
            delayed_count = 0
            tasks = db.query(models.Task).all()
            for t in tasks:
                if t.tien_do < 100.0 and t.budget:
                    b = t.budget
                    had_past_plan = (
                        b.thang_01_2025 > 0 or b.thang_02_2025 > 0 or
                        b.thang_03_2025 > 0 or b.thang_04_2025 > 0 or
                        b.quy_1_2026 > 0 or b.thang_05_2026 > 0 or
                        b.thang_06_2026 > 0
                    )
                    if had_past_plan:
                        delayed_count += 1

            # 2. Kiểm tra lệch ngân sách (> 10% lệch so với kế hoạch tháng gần nhất - Tháng 6/2026)
            plan_june_2026 = sum(t.budget.thang_06_2026 for t in tasks if t.budget)
            
            # Tính thực chi Tháng 6/2026
            spendings_june_2026 = db.query(models.ActualSpending).filter(
                models.ActualSpending.trang_thai_duyet == "Approved",
                models.ActualSpending.ngay_chi >= datetime.date(2026, 6, 1),
                models.ActualSpending.ngay_chi <= datetime.date(2026, 6, 30)
            ).all()
            actual_june_2026 = sum(s.so_tien_chi for s in spendings_june_2026)

            if plan_june_2026 > 0:
                deviation = abs(actual_june_2026 - plan_june_2026) / plan_june_2026
                if deviation > 0.10:
                    alert_msg = (
                        f"🚨 CẢNH BÁO LỆCH CHI TIÊU THÁNG 06/2026: Kế hoạch {plan_june_2026:,.2f} Trđ, "
                        f"Thực tế giải ngân {actual_june_2026:,.2f} Trđ (Lệch {deviation*100:.1f}% - vượt hạn mức 10%)."
                    )
                    print(f"[OBSERVABILITY] {alert_msg}")
                    await manager.broadcast({
                        "type": "red_alert",
                        "message": alert_msg,
                        "ma_ngan_sach": "THANG_06_2026"
                    })

            if delayed_count > 0:
                summary = f"⚠️ CẢNH BÁO TIẾN ĐỘ: Phát hiện {delayed_count} công việc trong quá khứ đã quá thời hạn phân bổ ngân sách nhưng chưa đạt tiến độ 100%."
                print(f"[OBSERVABILITY] {summary}")
                await manager.broadcast({
                    "type": "red_alert",
                    "message": summary,
                    "ma_ngan_sach": "TIEN_DO"
                })

        except Exception as e:
            print(f"Lỗi trong vòng lặp Observability: {e}")
        finally:
            db.close()
        
        await asyncio.sleep(20)  # Chạy quét sau mỗi 20 giây cho demo

@app.on_event("startup")
async def startup_event():
    asyncio.create_task(monitor_observability_loop())
